# Calculates cohen's d 
# Author: sudiptatung
# email: sudiptatung@gmail.com


import openpyxl as pxl
from math import sqrt
import ntpath

def cohensd():
	def ESize(k1,k2):
		m1,m2,s1,s2,n1,n2=sheet.cell(row=k1, column=2).value, sheet.cell(row=k2, column=2).value, sheet.cell(row=k1, column=3).value, sheet.cell(row=k2, column=3).value, sheet.cell(row=k1, column=4).value, sheet.cell(row=k2, column=4).value
		s_pooled=sqrt(((n2-1)*s2**2+(n1-1)*s1**2)/(n1+n2-2.0))
		return abs(m1-m2)/s_pooled
	def se(d1,j1,j2):
		n1,n2=sheet.cell(row=j1, column=4).value, sheet.cell(row=j2, column=4).value
		n=n1+n2
		return sqrt(((n-1.0)/(n-3.0))*(4.0/n)*(1+(d1**2)/8.0))
	#----------- output sheet ------------
	wb_out = pxl.Workbook()
	st_out = wb_out.active
	st_out.title = "Effect size results"
	ft = pxl.styles.Font(bold=True)
	count=1
	#--------------------------------------
	options=['y','yes','Y','Yes','YES']
	counter=0

	while counter==0:
		print '\nMention the path of the input excel file within inverted commas.\n'
		temp=0
		while temp==0:
			try:
				file_location=str(input('Input file path: '))
				wb=pxl.load_workbook(file_location)
				temp+=1
			except IOError:
				print '\nError: Write correct file path!\n'
			except SyntaxError:
				print '\nError: Write the file path within inverted comma!\n'
			
		file_name=ntpath.basename(file_location)
		st_out.cell(row=count, column=1).value ='Input file name: '+str(file_name)
		count+=1
		
		sheets=[str(i) for i in wb.get_sheet_names()]
		for i00 in range(len(sheets)):
			sheet=wb.get_sheet_by_name(sheets[i00])
			num_groups=sheet.max_row-1		
			
			st_out.cell(row=count, column=1).value = 'Datasheet: '+str(sheets[i00])##name of the sheet
			count+=1
			
			header_out=['Groups',"Cohen's d",'error(95%)']
			#row = st_out.row_dimensions[count]
			for ii in range(len(header_out)):
				st_out.cell(row=count, column=ii+1).value = header_out[ii]	
				st_out.cell(row=count, column=ii+1).font=ft			
			count+=1
			groups=[str(sheet.cell(row=i+2, column=1).value) for i in range(num_groups)]
			print 'Effect size for %s, %s:'%(str(file_name),sheets[i00])
			print 'Groups','\t',"Cohen's d",'\t','error(95%)'
			#st_out.write(count,0,
			for i1 in range(num_groups):
				for i2 in range(i1+1,num_groups):
					d=ESize(i1+2,i2+2)
					error=1.96*se(d,i1+2,i2+2)
					print '%s-%s'%(groups[i1],groups[i2]),'\t',d,'\t',error
					st_out.cell(row=count, column=1).value ='%s-%s'%(groups[i1],groups[i2])
					st_out.cell(row=count, column=2).value =d
					st_out.cell(row=count, column=3).value =error
					count+=1
			count+=1
			print ''
		count+=1
		exit_reply = raw_input("\nDo you have more input files for further effect size calculations? (y/n) ")
		if exit_reply not in options:
			save_input=raw_input("\nDo you want to create an output file? (y/n) ")		
			if options.count(save_input)>0:
				output_file_location=str(input("\nOutput file name: "))
				wb_out.save(output_file_location)
			counter+=1
			
if __name__ == "__main__":
	cohensd()
